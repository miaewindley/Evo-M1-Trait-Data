#!/usr/bin/env python3
"""
build_data.py -- regenerate __ShinyApp/data/ from the canonical repo files.

data/ is a BUILD OUTPUT, not hand-maintained. Re-run this whenever the merge
or trait tables change, and before deploying:

    python3 __ShinyApp/build_data.py

Single source of truth = the canonical files in the repo. Requires: openpyxl
(pip install openpyxl).
"""
import os, csv, re, shutil, urllib.parse, sys
import openpyxl

APP  = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.abspath(os.path.join(APP, ".."))
OUT  = os.path.join(APP, "data")
os.makedirs(OUT, exist_ok=True)

# ---- 1. fallback copies of the two compiled long tables ---------------------
# (Primary source at runtime is GitHub; these are only a local fallback.)
shutil.copyfile(os.path.join(REPO, "__merging_volumes", "volumes_long.csv"),
                os.path.join(OUT, "volumes_long.csv"))
shutil.copyfile(os.path.join(REPO, "__merging_cellcounts", "cellcounts_long.csv"),
                os.path.join(OUT, "cellcounts_long.csv"))

# ---- 2. melt the EvoM1 trait tables -> evom1_traits_long.csv -----------------
TT = os.path.join(REPO, "____EvoM1_TraitTable")
TRAIT_FILES = {
    "dexterity_corticospinaltract.xlsx": "Dexterity & corticospinal tract",
    "corticospinaltract_etc.xlsx":       "Corticospinal tract & ecology",
    "glia_gyrification.xlsx":            "Glia, gyrification & life history",
    "interlaminar_astrocytes.xlsx":      "Interlaminar astrocytes",
}
ID_COLS = {"species_sci", "Species", "Animal", "Species Generic Name"}
SRC_SUFFIX = ["_Source", " Source", "_Ref", " Ref", "_ref", " ref"]
is_src  = lambda c: any(c.endswith(s) for s in SRC_SUFFIX)
def base_of(c):
    for s in SRC_SUFFIX:
        if c.endswith(s):
            return c[:-len(s)].strip()
    return c

rows_out = []
for fn, label in TRAIT_FILES.items():
    ws = openpyxl.load_workbook(os.path.join(TT, fn), read_only=True, data_only=True)["Sheet1"]
    data = list(ws.iter_rows(values_only=True))
    hdr = [str(h).strip() if h is not None else "" for h in data[0]]
    src_idx = {base_of(c): i for i, c in enumerate(hdr) if is_src(c)}
    sp_i = hdr.index("Species") if "Species" in hdr else 0
    for r in data[1:]:
        sp = r[sp_i]
        if sp is None or not str(sp).strip() or str(sp).strip().lower() == "none":
            continue
        sp = str(sp).strip()
        for i, c in enumerate(hdr):
            if not c or c in ID_COLS or is_src(c):
                continue
            v = r[i]
            if v is None or not str(v).strip():
                continue
            vs = str(v).strip()
            if vs.lower() in ("na", "nan", "none", "-"):
                continue
            src = label
            if c in src_idx:
                sv = r[src_idx[c]]
                if sv is not None and str(sv).strip():
                    src = str(sv).strip()
            rows_out.append((sp, c, vs, src))

seen, dedup = set(), []
for row in rows_out:
    if row not in seen:
        seen.add(row); dedup.append(row)
with open(os.path.join(OUT, "evom1_traits_long.csv"), "w", newline="", encoding="utf-8") as fh:
    w = csv.writer(fh); w.writerow(["Species", "Variable", "Value", "Source"]); w.writerows(dedup)
print("evom1 traits rows:", len(dedup))

# ---- 3. index the public source tables (NOT copied) -------------------------
# The 180 source TSVs are served straight from __Public/comparative-data on
# GitHub at runtime, so they are not duplicated into the app. We only read them
# here to record row/column counts in the manifest.
PUB = os.path.join(REPO, "__Public", "comparative-data")
tsvs = sorted(f for f in os.listdir(PUB) if f.endswith(".tsv"))
mds  = [f for f in os.listdir(PUB) if f.endswith(".ReadMe.md")]
print("source tables indexed (served from GitHub):", len(tsvs))

# ---- 4. build source_manifest.csv (filenames + citations from __ReadMe.xlsx) -
wb = openpyxl.load_workbook(os.path.join(REPO, "__ReadMe.xlsx"), read_only=True, data_only=True)
ws = wb["Sheet1"]; rr = list(ws.iter_rows(values_only=True))
hdr = [str(h).strip() if h else "" for h in rr[0]]
C = {n: hdr.index(n) for n in ["Citation (APA 7th-Annotated)", "Item encoded", "1st Author", "year"]}
def g(r, n):
    v = r[C[n]]; return str(v).strip() if v is not None else ""
by_enc, by_pref = {}, {}
for r in rr[1:]:
    enc = g(r, "Item encoded")
    if enc:
        by_enc[enc] = r
        by_pref.setdefault(enc.split("_")[0], r)

def parse(fn):
    base = fn[:-4]
    ident_enc, label = (base.rsplit("_", 1) + [""])[:2] if "_" in base else (base, "")
    ident = urllib.parse.unquote(ident_enc)
    url, kind = "", "Other"
    if ident.startswith("10."):
        url, kind = "https://doi.org/" + ident, "DOI"
    elif ident.upper().startswith("PMID"):
        url, kind = "https://pubmed.ncbi.nlm.nih.gov/%s/" % re.sub(r"[^0-9]", "", ident), "PubMed"
    elif ident.upper().startswith("UMI"):
        kind = "Dissertation (ProQuest)"
    return ident_enc, ident, label, url, kind

manifest = []
for fn in tsvs:
    ident_enc, ident, label, url, kind = parse(fn)
    base = fn[:-4]
    r = by_enc.get(base) or by_pref.get(ident_enc)
    cit = g(r, "Citation (APA 7th-Annotated)") if r else ""
    auth = g(r, "1st Author") if r else ""
    yr = g(r, "year") if r else ""
    if auth and yr: short = "%s et al. (%s)" % (auth, yr)
    elif auth:      short = auth
    elif cit:       short = cit.split(".")[0][:60]
    else:           short = ident
    rm = ""
    for cand in (base + ".ReadMe.md", ident_enc + ".ReadMe.md"):
        if cand in mds:
            rm = cand; break
    with open(os.path.join(PUB, fn), newline="", encoding="utf-8", errors="replace") as fh:
        tab = list(csv.reader(fh, delimiter="\t"))
    ncol = len(tab[0]) if tab else 0
    hdrcols = "; ".join(h.strip().strip('"') for h in (tab[0] if tab else []))
    manifest.append(dict(file=fn, identifier=ident, id_type=kind, table_label=label,
                         url=url, readme=rm, n_rows=max(0, len(tab) - 1), n_cols=ncol,
                         columns=hdrcols, citation=cit, citation_short=short,
                         first_author=auth, year=yr))
fields = ["file", "identifier", "id_type", "table_label", "url", "readme",
          "n_rows", "n_cols", "columns", "citation", "citation_short",
          "first_author", "year"]
with open(os.path.join(OUT, "source_manifest.csv"), "w", newline="", encoding="utf-8") as fh:
    w = csv.DictWriter(fh, fieldnames=fields); w.writeheader(); w.writerows(manifest)
print("manifest rows:", len(manifest),
      "| with citation:", sum(1 for m in manifest if m["citation"]))
print("DONE ->", OUT)
